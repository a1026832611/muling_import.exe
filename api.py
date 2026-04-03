import base64
import time

import requests
import urllib3
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

AES_KEY = b"RmFeprFxZCmxyF72"
BASE_URL = "https://health.haiyy.cloud/api/c3"
SEX_MAP = {"男": "1", "女": "2"}
REQUEST_TIMEOUT = 30

# 角色对应的 ssrId
SSR_ID_MAP = {
    "睦邻-店员": "1050801353074540544",
    "睦邻-药师": "1050801314541469696",
    "睦邻-医生": "1050800673144307712",
}


def aes_ecb_encrypt(plain_text):
    """AES-ECB-PKCS7 加密，返回 Base64 字符串"""
    cipher = AES.new(AES_KEY, AES.MODE_ECB)
    padded = pad(plain_text.encode("utf-8"), AES.block_size)
    encrypted = cipher.encrypt(padded)
    return base64.b64encode(encrypted).decode("utf-8")


def _stringify_cell(value):
    """将 Excel 单元格值安全转成字符串，避免 None 变成字面量 'None'"""
    if value is None:
        return ""
    return str(value).strip()


def _parse_response_payload(resp):
    """解析接口响应，兼容空文本和非 JSON 响应"""
    text = resp.text or ""
    try:
        data = resp.json() if text else {}
    except ValueError:
        data = {}
    message = data.get("msg") or text[:100] or "接口未返回消息"
    return data, message


def login(account, password, img_code="8888"):
    """调用登录接口，返回 (token, client_id) 元组"""
    url = f"{BASE_URL}/login/login"
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
        "Origin": "https://csml.haiyy.cloud",
        "Referer": "https://csml.haiyy.cloud/",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/146.0.0.0 Safari/537.36",
    }
    data = {
        "account": account,
        "loginType": "1",
        "password": aes_ecb_encrypt(password),
        "imgCode": img_code,
    }
    resp = requests.post(url, data=data, headers=headers, verify=False, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    result = resp.json()
    user_data = result.get("data", {})
    token = user_data.get("apiToken")
    client_id = user_data.get("clientId")
    if not token:
        raise RuntimeError(f"登录失败: {result}")
    return token, client_id


def get_org_id(token, client_id):
    """获取机构ID，返回 (org_id, org_name)"""
    url = f"{BASE_URL}/organization/page"
    headers = {
        "accept": "application/json",
        "authorization": f"Bearer {token}",
        "clientid": client_id,
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/146.0.0.0 Safari/537.36",
    }
    params = {"size": 500, "clientId": client_id}
    resp = requests.get(url, headers=headers, params=params, verify=False, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    records = resp.json().get("data", {}).get("records", [])
    if not records:
        raise RuntimeError("未获取到机构数据")
    org_id = records[0]["id"]
    org_name = records[0].get("orgName", "")
    return org_id, org_name


def add_medical_care(name, phone, sex, id_card, token, client_id, add_org_id, ssr_id, password):
    """调用医疗护理添加接口"""
    url = f"{BASE_URL}/medicalCare/add"
    headers = {
        "accept": "application/json",
        "authorization": f"Bearer {token}",
        "clientid": client_id,
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/146.0.0.0 Safari/537.36",
    }
    params = {
        "account": phone,
        "clientId": client_id,
        "addOrgId": add_org_id,
        "name": name,
        "mobile": phone,
        "sex": sex,
        "cardType": "1",
        "cardNo": id_card,
        "password": password,
        "ssrId": ssr_id,
    }
    resp = requests.post(url, headers=headers, params=params, verify=False, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    return resp


def batch_add_from_excel(token, client_id, org_id, ssr_id, file_path, progress_callback=None):
    """从 Excel 批量读取数据并调用添加接口

    Args:
        progress_callback: 回调函数，签名为 callback(current, total, name, status, message)
            status: 'success' | 'fail' | 'skip'

    Returns:
        dict: {"success": int, "fail": int, "skip": int, "total": int}
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active
    total = max(sheet.max_row - 2, 0)  # 减去表头和模板行
    stats = {"success": 0, "fail": 0, "skip": 0, "total": total}

    for row in range(3, sheet.max_row + 1):
        current = row - 2
        name = _stringify_cell(sheet.cell(row=row, column=4).value)       # D列
        phone = _stringify_cell(sheet.cell(row=row, column=5).value)      # E列
        sex_value = _stringify_cell(sheet.cell(row=row, column=6).value)  # F列
        id_card = _stringify_cell(sheet.cell(row=row, column=7).value)    # G列
        password = _stringify_cell(sheet.cell(row=row, column=8).value)   # H列

        if not name or not id_card:
            stats["skip"] += 1
            if progress_callback:
                progress_callback(current, total, name or f"第{row}行", "skip", "数据不完整")
            continue

        sex = SEX_MAP.get(sex_value)
        if not sex:
            stats["skip"] += 1
            if progress_callback:
                progress_callback(current, total, name, "skip", f"性别值无效: {sex_value}")
            continue

        try:
            resp = add_medical_care(name, phone, sex, id_card, token, client_id, org_id, ssr_id, password)
            resp_data, msg = _parse_response_payload(resp)
            # 判断接口返回是否成功
            if resp_data.get("code") == 200 or resp_data.get("success"):
                stats["success"] += 1
                status = "success"
            else:
                stats["fail"] += 1
                status = "fail"
            if progress_callback:
                progress_callback(current, total, name, status, msg)
        except Exception as e:
            stats["fail"] += 1
            if progress_callback:
                progress_callback(current, total, name, "fail", str(e))

        time.sleep(0.5)

    workbook.close()
    return stats
